import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os

def calculate_age(birth_date_str):
    birth_date = datetime.fromisoformat(birth_date_str)
    today = datetime.today()
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    return age

def categorize_records(df):
    df['Вік'] = df['Дата народження'].apply(calculate_age)

    younger_18 = df[df['Вік'] < 18]
    age_18_45 = df[(df['Вік'] >= 18) & (df['Вік'] <= 45)]
    age_45_70 = df[(df['Вік'] > 45) & (df['Вік'] <= 70)]
    older_70 = df[df['Вік'] > 70]

    return {
        "all": df,
        "younger_18": younger_18,
        "18-45": age_18_45,
        "45-70": age_45_70,
        "older_70": older_70
    }

def write_xlsx(filename, sheets):
    try:
        wb = Workbook()
        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)
            # Форматування заголовків
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(filename)
        print("Дані успішно відформатовані!")
        print("Ok")
    except Exception as e:
        print(f"Помилка створення XLSX файлу: {e}")

def main():
    csv_file = 'employees.csv'
    xlsx_file = 'employees.xlsx'

    if not os.path.isfile(csv_file):
        print("Помилка відкриття файлу CSV.")
        return

    try:
        df = pd.read_csv(csv_file, sep=';', encoding='utf-8-sig')
    except Exception as e:
        print(f"Повідомлення про відсутність, або проблеми при відкритті файлу CSV: {e}")
        return

    sheets = categorize_records(df)
    write_xlsx(xlsx_file, sheets)

if __name__ == "__main__":
    main()
